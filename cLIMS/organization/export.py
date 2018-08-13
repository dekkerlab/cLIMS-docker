'''
Created on Nov 15, 2016

@author: nanda
'''
from django.contrib.auth.decorators import login_required
from django.http.response import HttpResponse, HttpResponseRedirect
from organization.models import *
import csv
import xlwt
from xlwt.compat import xrange
from openpyxl.reader.excel import load_workbook
from organization.excelRow import insert_rows
from openpyxl.styles import Color, Fill
from cLIMS.base import *
from dryLab.models import *
import json
from _collections import OrderedDict, defaultdict
from wetLab.models import *
import collections
from openpyxl.utils.cell import get_column_letter
from django.contrib import messages
import re


@login_required 
def exportExperiment(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="meta_data_sample.csv"'
    projectId = request.session['projectId']
    dbdata = Experiment.objects.filter(project=projectId)
    writer = csv.writer(response)
    for a in dbdata:
        field_names = [x.name for x in a._meta.local_fields]
        break
    writer.writerow(field_names)
    for obj in dbdata:
        writer.writerow([getattr(obj, field) for field in field_names])
    return response

def orderByNumber(jsonDict):
    jsonList = jsonDict.items()
    sorted_list = sorted(jsonList, key=lambda k: (int(k[1]['order'])))
    sorted_dict= OrderedDict(sorted_list)
    return sorted_dict

    
def export(analysisType,ws, projectId):
    dbdata = Analysis.objects.filter(analysis_exp__project=projectId, analysis_type__field_name=analysisType)
    if (dbdata):
        row_num = 0
        for a in dbdata:
            field_names = [x.name for x in a._meta.local_fields]
            break
        columns = []
        for names in field_names:
            if(names == "analysis_fields"):
                jsonObj = JsonObjField.objects.get(field_name=analysisType)
                jsonFields = orderByNumber(jsonObj.field_set)
                for keys in jsonFields:
                    columns.append((keys, 4000))
            elif(names == "analysis_import"):
                pass
            elif(names == "analysis_hiGlass"):
                pass
            else:
                columns.append((names, 4000))
        
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
    
        for col_num in xrange(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            # set column width
            ws.col(col_num).width = columns[col_num][1]
    
        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1
        
        for obj in dbdata:
            row_num += 1
            row = []
            for field in field_names:
                att = getattr(obj, field)
                if(field == "analysis_fields"):
                    json_data = json.loads(att)
                    for keys in jsonFields:
                        json_val = json_data[keys]
                        row.append(json_val)
                elif(field == "analysis_import"):
                    pass
                elif(field == "analysis_hiGlass"):
                    pass
                else:
                    if ((type(att) != int) and (type(att) != str)):
                        att = str(att)
                    row.append(att)
            for col_num in xrange(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)
    else:
        pass
    
@login_required 
def exportAnalysis(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    
    projectId = request.session['projectId']
    ana = Analysis.objects.filter(analysis_exp__project=projectId)
    
    analysis = list(set([x.analysis_type for x in ana]))
    
    for analysisType in analysis:
        print(analysisType, analysis)
        if str(analysisType) == "Hi-C Analysis":
            ws = wb.add_sheet('Hi-C')
            export(analysisType,ws, projectId)
            
        elif str(analysisType) == "3C Analysis":
            ws = wb.add_sheet('3C')
            export(analysisType,ws, projectId)
             
        elif str(analysisType) == "5C Analysis":
            ws = wb.add_sheet('5C')
            export(analysisType,ws, projectId)
        
        elif str(analysisType) == "CaptureC Analysis":
            ws = wb.add_sheet('CaptureC')
            export(analysisType,ws, projectId)
             
        else:
            pass
        
    wb.save(response)
    return response

@login_required 
def exportGEO(request):
    projectId = request.session['projectId']
    prj = Project.objects.get(pk=projectId)
    runUnits = SequencingRun.objects.filter(project=projectId)
    files = SeqencingFile.objects.filter(sequencingFile_exp__project=projectId).order_by('pk')
    experiments = Experiment.objects.filter(project=projectId)
    bioSample = Biosample.objects.filter(expBio__project=projectId)
    
    
    title = prj.project_name
    summary = prj.project_notes
    contributor1 = str(prj.project_owner)
    contributor2 = prj.project_contributor.all()
    membersList = []
    for values in contributor2:
        membersList.append(values)
     
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=GEO.xlsx'
    file_path_new = ROOTFOLDER+'/organization/static/siteWide/geo_template_new.xlsx'
 
    wb = load_workbook(file_path_new)
    ws = wb.worksheets[0]
    ws.insert_rows = insert_rows
     
    ws.cell(row=9, column=2).value = title
    ws.cell(row=10, column=2).value = summary
    ws.cell(row=12, column=2).value = contributor1
     
    memberRowNo = 13
    for members in membersList:
        insert_rows(ws, row_idx= memberRowNo, cnt = 1, above=True, copy_style=True)
        ws.cell(row=memberRowNo, column=1).value = "contributor"
        ws.cell(row=memberRowNo, column=2).value = str(members)
         
        memberRowNo +=1
     
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="Sample name"):
            sampleRowNo = i+2
            break
     
    count = 1
    
    for exp in experiments:
        insert_rows(ws, row_idx= sampleRowNo, cnt = 1, above=False, copy_style=False)
        ws.cell(row=sampleRowNo, column=1).value = "Sample " + str(count)
        ws.cell(row=sampleRowNo, column=2).value = str(exp.experiment_name)
        ws.cell(row=sampleRowNo, column=3).value = str(exp.experiment_biosample.biosample_biosource.biosource_tissue)
        ws.cell(row=sampleRowNo, column=4).value = str(exp.experiment_biosample.biosample_individual.individual_type)
        ws.cell(row=sampleRowNo, column=6).value = str(exp.experiment_enzyme.enzyme_name) 
        ws.cell(row=sampleRowNo, column=7).value = str(exp.experiment_biosample.biosample_biosource.biosource_cell_line)  
        ws.cell(row=sampleRowNo, column=8).value = str("DNA")
        ws.cell(row=sampleRowNo, column=9).value = str(exp.experiment_biosample.biosample_biosource.biosource_description)
        expFiles=SeqencingFile.objects.filter(sequencingFile_exp=exp)
        colC=11
        for f in expFiles:
            ws.cell(row=sampleRowNo, column=colC).value = str(f.sequencingFile_name)
            colC+=1
        sampleRowNo += 1
        count += 1
           
    rowNo = ws.max_row
    for i in range (rowNo):
        if((ws.cell(row=i+1, column=1).value)=="RAW FILES"):
            rawFilesRowNo = i+3
            break
     
    for file in files:
        insert_rows(ws, row_idx= rawFilesRowNo, cnt = 1, above=True, copy_style=False)
        ws.cell(row=rawFilesRowNo, column=1).value = str(file.sequencingFile_name)
        ws.cell(row=rawFilesRowNo, column=2).value = str(file.file_format)
        ws.cell(row=rawFilesRowNo, column=3).value = str(file.sequencingFile_sha256sum)
        ws.cell(row=rawFilesRowNo, column=4).value = str(file.sequencingFile_run.run_sequencing_instrument.choice_name)
        ws.cell(row=rawFilesRowNo, column=5).value = str(file.read_length)
        ws.cell(row=rawFilesRowNo, column=6).value = "paired-end"
        ws.cell(row=rawFilesRowNo, column=7).value = str(file.sequencingFile_mainPath)
        sampleRowNo += 1 
        rawFilesRowNo += 1
    
    #Find last row with data 
    for row in reversed(list(ws.rows)):
        valueData = [cell.value for cell in row]
        if any(valueData):
            pairedEndRowNo=int(row[0].row)+1
            break
        
    for file in files:
        if(file.related_files):
            ws.cell(row=pairedEndRowNo, column=1).value = str(file.related_files.sequencingFile_name)
            ws.cell(row=pairedEndRowNo, column=2).value = str(file.sequencingFile_name)
            pairedEndRowNo+=1
    
    
    wb.save(response)
    return response

def appendFilterdcic(dcicExcelSheet,sheetname,entry):
    if(not(entry[0].split(":")[0]=="dcic")):
        dcicExcelSheet[sheetname].append(entry)
        
        
def update_dcic(obj):
    obj.update_dcic = False
    obj.save()
                
def initialize(tab,sheetTab):
    file_path_new = ROOTFOLDER+'/organization/static/siteWide/DCIC-Metadata_entry_form_FEB2018.xlsx'
    wb = load_workbook(file_path_new)
    sheet = wb.get_sheet_by_name(tab)
    maxCol=get_column_letter(sheet.max_column)
    headers = []
    
    for rowOfCellObjects in sheet['B1':maxCol]:         
        for cellObj in rowOfCellObjects:
            if "*" in cellObj.value:
                headers.append(cellObj.value[1:])
            else:
                headers.append(cellObj.value)
    
    sheetTab.append(headers)
    return (sheetTab)

def appendingFunc(columnNamesList,obj,indx,appString):
    obj[columnNamesList.index(indx)]=str(appString)

def appendLab(objClass,sheetExcel,singleItem,dcicExcelSheet):
    if(objClass.contributing_labs.all()):
        contrilabs=[]
        for labs in objClass.contributing_labs.all():
            contrilabs.append(labs.lab_name)
        appendingFunc(dcicExcelSheet[sheetExcel][0],singleItem,"contributing_labs",(",".join(contrilabs)))

def appendPublication(pKey, dcicExcelSheet,finalizeOnly):
    pub = Publication.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(pub)
    singlePub = [""] * len(dcicExcelSheet["Publication"][0])
    appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"aliases",(pub.dcic_alias))
    appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"title",(pub.publication_title))
    appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"ID",(pub.publication_id))
    if(pub.attachment):
        appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"attachment",(str(FILEUPLOADPATH)+str(pub.attachment)))
    if(pub.publication_categories):
        appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"categories",(pub.publication_categories))
    appendLab(pub,"Publication",singlePub,dcicExcelSheet)
    if(pub.exp_sets_prod_in_pub):
        appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"exp_sets_prod_in_pub",(pub.exp_sets_prod_in_pub))
    if(pub.exp_sets_used_in_pub):
        appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"exp_sets_used_in_pub",(pub.exp_sets_used_in_pub))
    if(pub.publication_published_by):
        appendingFunc(dcicExcelSheet["Publication"][0],singlePub,"published_by",(pub.publication_published_by))
    appendFilterdcic(dcicExcelSheet,'Publication',singlePub)

def appendDocument(pKey, dcicExcelSheet,finalizeOnly):
    doc = Document.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(doc)
    singleDocument = [""] * len(dcicExcelSheet["Document"][0])
    appendingFunc(dcicExcelSheet["Document"][0],singleDocument,"aliases",(doc.dcic_alias))
    appendingFunc(dcicExcelSheet["Document"][0],singleDocument,"description",(doc.description))
    if(doc.attachment):
        appendingFunc(dcicExcelSheet["Document"][0],singleDocument,"description",(str(FILEUPLOADPATH)+str(doc.attachment)))
    appendLab(doc,"Document",singleDocument,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["Document"][0],singleDocument,"urls",(doc.url))
    if(doc.references):
        appendingFunc(dcicExcelSheet["Document"][0],singleDocument,"references",(doc.references.dcic_alias))
        appendPublication(doc.references.pk,dcicExcelSheet,finalizeOnly)
    appendFilterdcic(dcicExcelSheet,'Document',singleDocument)

def appendVendor(pKey,dcicExcelSheet,finalizeOnly):
    ven = Vendor.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(ven)
    singleVendor = [""] * len(dcicExcelSheet["Vendor"][0])
    appendingFunc(dcicExcelSheet["Vendor"][0],singleVendor,"aliases",(ven.dcic_alias))
    appendingFunc(dcicExcelSheet["Vendor"][0],singleVendor,"title",(ven.vendor_title))
    if(ven.vendor_description != None):
        appendingFunc(dcicExcelSheet["Vendor"][0],singleVendor,"description",(ven.vendor_description))
    appendingFunc(dcicExcelSheet["Vendor"][0],singleVendor,"url",(ven.vendor_url))
    appendFilterdcic(dcicExcelSheet,'Vendor',singleVendor)

def appendEnzyme(pKey,dcicExcelSheet,finalizeOnly):
    enz = Enzyme.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(enz)
    singleEnzyme = [""] * len(dcicExcelSheet["Enzyme"][0])
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"aliases",(enz.dcic_alias))
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"name",(enz.enzyme_name))
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"description",(enz.enzyme_description))
    if(enz.enzyme_vendor):
        appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"enzyme_source",(enz.enzyme_vendor.dcic_alias))
        appendVendor(enz.enzyme_vendor.pk, dcicExcelSheet,finalizeOnly)
    appendLab(enz,"Enzyme",singleEnzyme,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"catalog_number",(enz.enzyme_catalog_number))
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"recognition_sequence",(enz.enzyme_recogSeq))
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"site_length",(enz.enzyme_siteLen))
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"cut_position",(enz.enzyme_cutPos))
    if(enz.document):
        appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"documents",(enz.document.dcic_alias))
        appendDocument(enz.document.pk, dcicExcelSheet,finalizeOnly)
    appendingFunc(dcicExcelSheet["Enzyme"][0],singleEnzyme,"url",(enz.url))
    appendFilterdcic(dcicExcelSheet,'Enzyme',singleEnzyme)

def appendAntibody(pKey,dcicExcelSheet,finalizeOnly):
    anti=Antibody.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(anti)
    singleAntibody = [""] * len(dcicExcelSheet["Antibody"][0])
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"aliases",(anti.dcic_alias))
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_name",(anti.antibody_name))
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_product_no",(anti.antibody_product_no))
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"description",(anti.description))
    
    if(anti.antibody_target):
        appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_target",(anti.antibody_target.dcic_alias))
        appendTarget(anti.antibody_target.pk, dcicExcelSheet,finalizeOnly)
    
    if(anti.antibody_vendor):
        appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_vendor",(anti.antibody_vendor.dcic_alias))
        appendVendor(anti.antibody_vendor.pk, dcicExcelSheet,finalizeOnly)
    
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_vendor",(anti.antibody_vendor))
    appendingFunc(dcicExcelSheet["Antibody"][0],singleAntibody,"antibody_encode_accession",(anti.antibody_encode_accession))
    

def appendImageObjects(pKey,dcicExcelSheet,finalizeOnly):
    img=ImageObjects.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(img)
    singleItem = [""] * len(dcicExcelSheet["Image"][0])
    appendingFunc(dcicExcelSheet["Image"][0],singleItem,"aliases",(img.dcic_alias))
    appendingFunc(dcicExcelSheet["Image"][0],singleItem,"attachment",(str(FILEUPLOADPATH)+str(img.imageObjects_images)))
    appendLab(img,"Image",singleItem,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["Image"][0],singleItem,"caption",(img.description))
    appendFilterdcic(dcicExcelSheet,'Image',singleItem)
    
def appendConstruct(pKey,dcicExcelSheet,finalizeOnly):
    construct = Construct.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(construct)
    singleItem = [""] * len(dcicExcelSheet["Construct"][0])
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"aliases",(construct.dcic_alias))
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"name",(construct.construct_name))
    if(construct.construct_description != None):
        appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"description",(construct.construct_description))
    if(construct.construct_type != None):
        appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"construct_type",(construct.construct_type))
    if(construct.construct_vendor):
        appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"construct_vendor",(construct.construct_vendor.dcic_alias))
        appendVendor(construct.construct_vendor.pk,dcicExcelSheet,finalizeOnly)
    appendLab(construct,"Construct",singleItem,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"designed_to_target",(construct.construct_designed_to_Target))
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"insert_sequence",(construct.construct_insert_sequence))
    if(construct.construct_map):
        appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"map",(construct.construct_map.dcic_alias))
        appendDocument(construct.construct_map.pk, dcicExcelSheet,finalizeOnly)
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"protein_tags",(construct.construct_tag))
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"vector_backbone",(construct.construct_vector_backbone))
    if(construct.references):
        appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"references",(construct.references.dcic_alias))
    appendingFunc(dcicExcelSheet["Construct"][0],singleItem,"url",(construct.url))
    appendFilterdcic(dcicExcelSheet,'Construct',singleItem)

def appendTarget(pKey,dcicExcelSheet,finalizeOnly):
    target=Target.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(target)
    singleItem = [""] * len(dcicExcelSheet["Target"][0])
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"aliases",(target.dcic_alias))
    if(target.target_description != None):
        appendingFunc(dcicExcelSheet["Target"][0],singleItem,"description",(target.target_description))
    appendLab(target,"Target",singleItem,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"targeted_genes",(target.targeted_genes))
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"targeted_genome_regions",(target.targeted_region))
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"targeted_proteins",(target.targeted_proteins))
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"targeted_rnas",(target.targeted_rnas))
    if(target.targeted_structure!=None):
        appendingFunc(dcicExcelSheet["Target"][0],singleItem,"targeted_structure",(target.targeted_structure))
    if(target.references):
        appendingFunc(dcicExcelSheet["Target"][0],singleItem,"references",(target.references.dcic_alias))
        appendPublication(target.references.pk,dcicExcelSheet,finalizeOnly)
    appendingFunc(dcicExcelSheet["Target"][0],singleItem,"dbxrefs",(target.dbxrefs))
    appendFilterdcic(dcicExcelSheet,'Target',singleItem)
    
def appendGenomicRegion(pKey,dcicExcelSheet,finalizeOnly):          
    genomicRegion = GenomicRegions.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(genomicRegion)
    singleItem = [""] * len(dcicExcelSheet["GenomicRegion"][0])
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"aliases",(genomicRegion.dcic_alias))
    if(genomicRegion.genomicRegions_genome_assembly):
        appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"genome_assembly",(genomicRegion.genomicRegions_genome_assembly))
    if(genomicRegion.genomicRegions_chromosome != None):
        appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"chromosome",(genomicRegion.genomicRegions_chromosome))
    appendLab(genomicRegion,"GenomicRegion",singleItem,dcicExcelSheet)
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"start_coordinate",(genomicRegion.genomicRegions_start_coordinate))
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"end_coordinate",(genomicRegion.genomicRegions_end_coordinate))
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"location_description",(genomicRegion.genomicRegions_location_description))
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"start_location",(genomicRegion.genomicRegions_start_location))
    appendingFunc(dcicExcelSheet["GenomicRegion"][0],singleItem,"end_location",(genomicRegion.genomicRegions_end_location))
    appendFilterdcic(dcicExcelSheet,'GenomicRegion',singleItem)

def appendModification(pKey,dcicExcelSheet,finalizeOnly):
    modificationObj = Modification.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(modificationObj)
    singleMod = [""] * len(dcicExcelSheet["Modification"][0])
    appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"aliases",(modificationObj.dcic_alias))
    appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"description",(modificationObj.modification_description))
    if(str(modificationObj.modification_type) != None):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"modification_type",(modificationObj.modification_type))
    if(modificationObj.constructs):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"constructs",(modificationObj.constructs.dcic_alias))
        appendConstruct(modificationObj.constructs.pk, dcicExcelSheet,finalizeOnly)
    appendLab(modificationObj,"Modification",singleMod,dcicExcelSheet)
    if(modificationObj.modification_vendor):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"created_by",(modificationObj.modification_vendor.dcic_alias))
        appendVendor(modificationObj.modification_vendor.pk, dcicExcelSheet,finalizeOnly)
    appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"guide_rnas",(modificationObj.modification_gRNA))
    if(modificationObj.modification_genomicRegions):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"modified_regions",(modificationObj.modification_genomicRegions.dcic_alias))
        appendGenomicRegion(modificationObj.modification_genomicRegions.pk, dcicExcelSheet,finalizeOnly)
    if(modificationObj.target):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"target_of_mod",(modificationObj.target.dcic_alias))
        appendTarget(modificationObj.target.pk, dcicExcelSheet,finalizeOnly)
    if(modificationObj.references):
        appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"references",(modificationObj.references.dcic_alias))
        appendPublication(modificationObj.references.pk,dcicExcelSheet,finalizeOnly)
    appendingFunc(dcicExcelSheet["Modification"][0],singleMod,"url",(modificationObj.url))
    appendFilterdcic(dcicExcelSheet,'Modification',singleMod)
    


def appendProtocol(pKey,dcicExcelSheet,finalizeOnly):
    protocolObj = Protocol.objects.get(pk=pKey)
    if(finalizeOnly):
        update_dcic(protocolObj)
    singleProtocol = [""] * len(dcicExcelSheet["Protocol"][0])
    appendingFunc(dcicExcelSheet["Protocol"][0],singleProtocol,"aliases",(protocolObj.dcic_alias))
    appendingFunc(dcicExcelSheet["Protocol"][0],singleProtocol,"description",(protocolObj.description))
    if(protocolObj.protocol_type != None):
        appendingFunc(dcicExcelSheet["Protocol"][0],singleProtocol,"protocol_type",(protocolObj.protocol_type.choice_name))
    if(protocolObj.attachment):
        appendingFunc(dcicExcelSheet["Protocol"][0],singleProtocol,"attachment",(str(FILEUPLOADPATH)+str(protocolObj.attachment)))
    appendLab(protocolObj,"Protocol",singleProtocol,dcicExcelSheet)
    if(protocolObj.protocol_classification != None):
        appendingFunc(dcicExcelSheet["Protocol"][0],singleProtocol,"protocol_classification",(protocolObj.protocol_classification.choice_name))
    appendFilterdcic(dcicExcelSheet,'Protocol',singleProtocol)


def appendFiles(pKey,dcicExcelSheet,finalizeOnly):
    f = SeqencingFile.objects.get(pk=pKey)
    if(str(f.file_format)=="fasta"):
        singleFile = [""] * len(dcicExcelSheet["FileFasta"][0])
        appendingFunc(dcicExcelSheet["FileFasta"][0],singleFile,"aliases",(f.dcic_alias))
        appendingFunc(dcicExcelSheet["FileFasta"][0],singleFile,"file_format",(f.file_format))
        appendLab(f,"FileFasta",singleFile,dcicExcelSheet)
        if(f.file_format_specifications):
            appendingFunc(dcicExcelSheet["FileFasta"][0],singleFile,"file_format_specifications",(f.file_format_specifications))
        appendingFunc(dcicExcelSheet["FileFasta"][0],singleFile,"dbxrefs",(f.dbxrefs))
        ##FileMainPATH is exported then manually remove it so that we know the path
        singleFile.append(f.sequencingFile_mainPath)
        
        appendFilterdcic(dcicExcelSheet,'FileFasta',singleFile)
        
    elif(str(f.file_format)=="fastq"):
        singleFile = [""] * len(dcicExcelSheet["FileFastq"][0])
        appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"aliases",(f.dcic_alias))
        appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"file_format",(f.file_format))
        appendLab(f,"FileFastq",singleFile,dcicExcelSheet)
        if(f.file_format_specifications):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"file_format_specifications",(f.file_format_specifications))
        if(f.file_barcode):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.barcode",(f.file_barcode.barcode_index))
        if(f.barcode_in_read):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.barcode_in_read",(f.barcode_in_read))
        if(f.file_barcode):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.barcode_position",(f.file_barcode.barcode_position))
        if(f.flowcell_details_chunk != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.chunk",(f.flowcell_details_chunk))
        appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.lane",(f.flowcell_details_lane))
        if(f.sequencingFile_run.run_sequencing_machine != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"flowcell_details.machine",(f.sequencingFile_run.run_sequencing_machine))
        if(f.sequencingFile_run.run_sequencing_instrument != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"instrument",(f.sequencingFile_run.run_sequencing_instrument))
        if(f.paired_end != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"paired_end",(f.paired_end))
        if(f.read_length != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"read_length",(f.read_length))
        if(f.relationship_type != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"related_files.relationship_type",(f.relationship_type))
        if(f.related_files != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"related_files.file",(f.related_files.dcic_alias))
        if(f.dbxrefs != None):
            appendingFunc(dcicExcelSheet["FileFastq"][0],singleFile,"dbxrefs",(f.dbxrefs))
        ##FileMainPATH is exported then manually remove it so that we know the path
        singleFile.append(f.sequencingFile_mainPath)
        appendFilterdcic(dcicExcelSheet,'FileFastq',singleFile)

def appendBioRep(expPk,singleExp):
    exp = Experiment.objects.get(pk=expPk)
    expSameBiosource = Experiment.objects.filter(experiment_biosample__biosample_biosource=exp.experiment_biosample.biosample_biosource,
                                                protocol=exp.protocol,experiment_enzyme=exp.experiment_enzyme, 
                                                type=exp.type,project=exp.project)

    
    biosampleFields=json.loads(exp.experiment_biosample.biosample_fields)
    bioReplicates = []
    #biosamPk = []
    fieldsToCheckBiosample=["synchronization_stage","karyotype"]
    
    for e in expSameBiosource:
        #biosamPk.append(e.experiment_biosample.pk)
        #bioReplicates=list(set(biosamPk))
        sameFieldsBiosample=json.loads(e.experiment_biosample.biosample_fields)
        #if((sorted(expSameFields.items()) == sorted(expFields.items()))):
        if(exp.type.field_name in ["Hi-C Exp Protocol","CaptureC Exp Protocol","ATAC-seq Protocol", "3C Exp Protocol"]):
            if( all(biosampleFields[x] == sameFieldsBiosample[x] for x in fieldsToCheckBiosample) 
                and set(e.experiment_biosample.modifications.all())==set(exp.experiment_biosample.modifications.all()) 
                and e.experiment_biosample.protocol==exp.experiment_biosample.protocol
                and set(e.experiment_biosample.biosample_TreatmentRnai.all())==set(exp.experiment_biosample.biosample_TreatmentRnai.all())
                and set(e.experiment_biosample.biosample_TreatmentChemical.all())==set(exp.experiment_biosample.biosample_TreatmentChemical.all())):
                bioReplicates.append(e.experiment_biosample.pk)
        else:
            bioReplicates.append(0)

    bio_rep_no = (sorted(bioReplicates)).index(exp.experiment_biosample.pk)+1
    singleExp.append(bio_rep_no)
    
def appendTechRep(expPk,singleExp):
    exp = Experiment.objects.get(pk=expPk)
    expSameBiosample = Experiment.objects.filter(experiment_biosample=exp.experiment_biosample,project=exp.project)
    
#     expFields=json.loads(exp.experiment_fields)
    techReplicates = []
#     fieldsToCheckExpHicCaptureC=["crosslinking_time","experiment_type","average_fragment_size","digestion_temperature",
#                                  "ligation_time","digestion_time","tagging_method","ligation_volume","crosslinking_method",
#                                  "fragmentation_method","ligation_temperature","crosslinking_temperature","biotin_removed",
#                                  "fragment_size_range"]
    for e in expSameBiosample:
#         expSameFields=json.loads(e.experiment_fields)
#         if( all(expSameFields[x] == expFields[x] for x in fieldsToCheckExpHicCaptureC)):
        techReplicates.append(e.pk)
    
    tech_rep_no = (sorted(techReplicates)).index(expPk)+1
    
    singleExp.append(tech_rep_no)
    
    
def populateDict(request, experimentList):
    finalizeOnly = request.session['finalizeOnly']
    projectId = request.session['projectId']
    bioSample = Biosample.objects.filter(expBio__pk__in=experimentList)
    dcicExcelSheet=defaultdict(list)

    tabNames = ("Document","Protocol","Publication","IndividualMouse","IndividualHuman","Vendor","Enzyme","Antibody","Biosource","Construct","TreatmentRnai",
                "TreatmentChemical","GenomicRegion","Target","Modification","Image","BiosampleCellCulture","Biosample","FileFastq","FileFasta",
                "ExperimentHiC","ExperimentCaptureC","ExperimentAtacseq","ExperimentSeq","ExperimentSetReplicate","ExperimentSet")
    
    for tab in tabNames:
        dcicExcelSheet[tab] = initialize(tab, dcicExcelSheet[tab])
    
    ##Biosample
    for sample in bioSample:
        if(finalizeOnly):
            update_dcic(sample)
        singleSample = []
        singleSample=[""] * len(dcicExcelSheet["Biosample"][0])
        appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"aliases",(sample.dcic_alias))
        appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"description",(sample.biosample_description))
        appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"biosource",(sample.biosample_biosource.dcic_alias))
        if(sample.protocol):
            appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"biosample_protocols",(sample.protocol.dcic_alias))
            appendProtocol(sample.protocol.pk,dcicExcelSheet,finalizeOnly)
        if(sample.biosample_type):
            ##BiosampleCellCulture is a special case for DCIC Alias since there is only one object Biosample for this
            aliasList=sample.dcic_alias.split("_")
            appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"cell_culture_details",(LABNAME +"BiosampleCellCulture_"+str("_".join(aliasList[1:]))))
        appendLab(sample,"Biosample",singleSample,dcicExcelSheet)
        if(sample.modifications.all()):
            modList = []
            for mod in sample.modifications.all():
                if(finalizeOnly):
                    update_dcic(mod)
                modList.append(mod.dcic_alias)
                appendModification(mod.pk,dcicExcelSheet,finalizeOnly)
            appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"modifications",(",".join(modList)))
        rnai= []
        chemical=[]
        treatmentList=[]
        if(sample.biosample_TreatmentRnai.all()):
            for rnaiTreat in sample.biosample_TreatmentRnai.all():
                if(finalizeOnly):
                    update_dcic(rnaiTreat)
                rnai.append(rnaiTreat.dcic_alias)
            treatmentList.append(",".join(rnai))
        if(sample.biosample_TreatmentChemical.all()):
            for chemTreat in sample.biosample_TreatmentChemical.all():
                if(finalizeOnly):
                    update_dcic(chemTreat)
                chemical.append(chemTreat.dcic_alias)
            treatmentList.append(",".join(chemical))
        appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"treatments",(",".join(treatmentList)))
        
        if(sample.references):
            appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"references",(sample.references.dcic_alias))
            appendPublication(sample.references.pk,dcicExcelSheet,finalizeOnly)
        
        if(sample.dbxrefs):
            appendingFunc(dcicExcelSheet["Biosample"][0],singleSample,"dbxrefs",(sample.dbxrefs))
        
        appendFilterdcic(dcicExcelSheet,'Biosample',singleSample)
        

        ##Biosamplecellculture
        if(sample.biosample_type):
            bcc=json.loads(sample.biosample_fields)
            singleBcc = [""] * len(dcicExcelSheet["BiosampleCellCulture"][0])
            ##BiosampleCellCulture is a special case for DCIC Alias since there is only one object Biosample for this
            aliasList=sample.dcic_alias.split("_")
            
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"aliases",(LABNAME +"BiosampleCellCulture_"+str("_".join(aliasList[1:]))))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"description",(sample.biosample_description))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"culture_start_date",(bcc["culture_start_date"]))
            
            
            ##authentication_protocols
            authDocs=[]
            if(sample.authentication_protocols.all()):
                for authDoc in sample.authentication_protocols.all():
                    if(finalizeOnly):
                        update_dcic(authDoc)
                    authDocs.append(authDoc.dcic_alias)
                    appendProtocol(authDoc.pk,dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"authentication_protocols",(",".join(authDocs)))
            
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"cell_line_lot_number",(bcc["cell_line_lot_number"]))
            appendLab(sample,"BiosampleCellCulture",singleBcc,dcicExcelSheet)
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"culture_duration",(bcc["culture_duration"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"culture_harvest_date",(bcc["culture_harvest_date"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"differentiation_state",(bcc["differentiation_state"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"doubling_number",(bcc["doubling_number"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"doubling_time",(bcc["doubling_time"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"follows_sop",(bcc["follows_sop"]))
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"karyotype",(bcc["karyotype"]))

            if(ImageObjects.objects.filter(bioImg__pk=sample.pk)):
                image=ImageObjects.objects.filter(bioImg__pk=sample.pk)
                ig1 =[ imgs for imgs in image if imgs.imageObjects_type.choice_name=="morphology_image"]
                if(ig1):
                    appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"morphology_image",(ig1[0].dcic_alias))
                    appendImageObjects(ig1[0].pk,dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"passage_number",(bcc["passage_number"]))
            
#             singleBcc.append("") ##protocol_SOP_deviations
                ##protocol_additional
            if(Protocol.objects.filter(bioAddProto__pk=sample.pk)):
                proto = Protocol.objects.filter(bioAddProto__pk=sample.pk)
                if(proto.all()):
                    protoList = []
                    for p in proto.all():
                        if(finalizeOnly):
                            update_dcic(p)
                        protoList.append(p.dcic_alias)
                        appendProtocol(p.pk,dcicExcelSheet,finalizeOnly)
                    appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"protocols_additional",(",".join(protoList)))
            
            appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"synchronization_stage",(bcc["synchronization_stage"]))
            
            if(sample.dbxrefs != None ):
                appendingFunc(dcicExcelSheet["BiosampleCellCulture"][0],singleBcc,"dbxrefs",(sample.dbxrefs))

            appendFilterdcic(dcicExcelSheet,'BiosampleCellCulture',singleBcc)
            
        ##treatments
        if(sample.biosample_TreatmentRnai):
            treatmentRnais = TreatmentRnai.objects.filter(biosamTreatmentRnai=sample.pk)
            for treatmentRnai in treatmentRnais:
                if(finalizeOnly):
                    update_dcic(treatmentRnai)
                singleItem = [""] * len(dcicExcelSheet["TreatmentRnai"][0])
                appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"aliases",(treatmentRnai.dcic_alias))
                appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"description",(treatmentRnai.treatmentRnai_description))
                if(treatmentRnai.treatmentRnai_type):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"rnai_type",(treatmentRnai.treatmentRnai_type))
                appendLab(treatmentRnai,"TreatmentRnai",singleItem,dcicExcelSheet)
                if(treatmentRnai.constructs):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"rnai_constructs",(treatmentRnai.constructs.dcic_alias))
                    appendConstruct(treatmentRnai.constructs.pk, dcicExcelSheet,finalizeOnly)
                if(treatmentRnai.treatmentRnai_vendor):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"rnai_vendor",(treatmentRnai.treatmentRnai_vendor.dcic_alias))
                    appendVendor(treatmentRnai.treatmentRnai_vendor.pk, dcicExcelSheet,finalizeOnly)
                if(treatmentRnai.treatmentRnai_target):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"target",(treatmentRnai.treatmentRnai_target.dcic_alias))
                    appendTarget(treatmentRnai.treatmentRnai_target.pk, dcicExcelSheet,finalizeOnly)
                if(treatmentRnai.treatmentRnai_nucleotide_seq):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"target_sequence",(treatmentRnai.treatmentRnai_nucleotide_seq))
                if(treatmentRnai.document):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"documents",(treatmentRnai.document.dcic_alias))
                    appendDocument(treatmentRnai.document.pk, dcicExcelSheet,finalizeOnly)
                if(treatmentRnai.references):
                    appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"references",(treatmentRnai.references.dcic_alias))
                    appendPublication(treatmentRnai.references.pk,dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["TreatmentRnai"][0],singleItem,"url",(treatmentRnai.url))
                appendFilterdcic(dcicExcelSheet,'TreatmentRnai',singleItem)
                
        
        if(sample.biosample_TreatmentChemical):
            treatmentChemicals = TreatmentChemical.objects.filter(biosamTreatmentChemical=sample.pk)
            for treatmentChemical in treatmentChemicals:
                if(finalizeOnly):
                    update_dcic(treatmentChemical)
                singleItem = [""] * len(dcicExcelSheet["TreatmentChemical"][0])
                appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"aliases",(treatmentChemical.dcic_alias))
                appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"description",(treatmentChemical.treatmentChemical_description))
                appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"chemical",(treatmentChemical.treatmentChemical_chemical))
                if(treatmentChemical.treatmentChemical_concentration != 0):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"concentration",(treatmentChemical.treatmentChemical_concentration))
                if(treatmentChemical.treatmentChemical_concentration_units != None):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"concentration_units",(treatmentChemical.treatmentChemical_concentration_units))
                appendLab(treatmentChemical,"TreatmentChemical",singleItem,dcicExcelSheet)
                if(treatmentChemical.treatmentChemical_duration != None):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"duration",(treatmentChemical.treatmentChemical_duration))
                if(treatmentChemical.treatmentChemical_duration_units != None):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"duration_units",(treatmentChemical.treatmentChemical_duration_units))
                if(treatmentChemical.treatmentChemical_temperature != None):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"temperature",(treatmentChemical.treatmentChemical_temperature))
                if(treatmentChemical.document):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"documents",(treatmentChemical.document.dcic_alias))
                    appendDocument(treatmentChemical.document.pk, dcicExcelSheet,finalizeOnly)
                if(treatmentChemical.references):
                    appendingFunc(dcicExcelSheet["TreatmentChemical"][0],singleItem,"references",(treatmentChemical.references.dcic_alias))
                    appendPublication(treatmentChemical.references.pk,dcicExcelSheet,finalizeOnly)
                appendFilterdcic(dcicExcelSheet,'TreatmentChemical',singleItem)
                
        if(Biosource.objects.get(bioSource__pk=sample.pk)):
            biosource = Biosource.objects.get(bioSource__pk=sample.pk)
            if(finalizeOnly):
                update_dcic(biosource)
            singleBio = [""] * len(dcicExcelSheet["Biosource"][0])
            appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"aliases",(biosource.dcic_alias))
            appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"description",(biosource.biosource_description))
            appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"biosource_type",(biosource.biosource_type))
            appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"cell_line",(biosource.biosource_cell_line))
               
            ###Standard operation protocol 
            if(biosource.protocol):
                appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"SOP_cell_line",(biosource.protocol.dcic_alias))
                appendProtocol(biosource.protocol.pk, dcicExcelSheet, finalizeOnly)
            if(biosource.biosource_vendor):
                appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"biosource_vendor",(biosource.biosource_vendor.dcic_alias))
                appendVendor(biosource.biosource_vendor.pk, dcicExcelSheet,finalizeOnly)
            appendLab(biosource,"Biosource",singleBio,dcicExcelSheet)        
            if(biosource.biosource_individual):
                appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"individual",(biosource.biosource_individual.dcic_alias))
                indi = Individual.objects.get(pk=biosource.biosource_individual.pk)
                if(finalizeOnly):
                    update_dcic(indi)
                indiJson = json.loads(indi.individual_fields)
                singleIndi = [""] * len(dcicExcelSheet["IndividualMouse"][0])
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"aliases",(indi.dcic_alias))
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"age",(indiJson["age"]))
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"age_units",(indiJson["age_units"]))
                appendLab(biosource.biosource_individual,"IndividualMouse",singleIndi,dcicExcelSheet)  
                if(str(biosource.biosource_individual.individual_type)=="IndividualMouse"):
                    appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"mouse_life_stage",(indiJson["mouse_life_stage"]))
                    appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"mouse_life_stage",(indiJson["mouse_strain"]))
                    if(indi.individual_vendor):
                        appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"mouse_vendor",(indi.individual_vendor.dcic_alias))
                        appendVendor(indi.individual_vendor.pk, dcicExcelSheet,finalizeOnly)
             
                elif(str(biosource.biosource_individual.individual_type)=="IndividualHuman"):
                    appendingFunc(dcicExcelSheet["IndividualHuman"][0],singleIndi,"ethnicity",(indiJson["ethnicity"]))
                    appendingFunc(dcicExcelSheet["IndividualHuman"][0],singleIndi,"health_status",(indiJson["health_status"]))
                    appendingFunc(dcicExcelSheet["IndividualHuman"][0],singleIndi,"life_stage",(indiJson["life_stage"]))
                
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"sex",(indiJson["sex"]))
                
                if(indi.document):
                    appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"documents",(indi.document.dcic_alias))
                    appendDocument(indi.document.pk, dcicExcelSheet,finalizeOnly)
                
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"url",(indi.url))
                appendingFunc(dcicExcelSheet["IndividualMouse"][0],singleIndi,"dbxrefs",(indi.dbxrefs))
                 
                if(str(biosource.biosource_individual.individual_type)=="IndividualMouse"):
                    appendFilterdcic(dcicExcelSheet,'IndividualMouse',singleIndi)
                    
                if(str(biosource.biosource_individual.individual_type)=="IndividualHuman"):
                    appendFilterdcic(dcicExcelSheet,'IndividualHuman',singleIndi)
                   
                
            if(biosource.modifications):
                modList = []
                for mod in biosource.modifications.all():
                    if(finalizeOnly):
                        update_dcic(mod)
                    modList.append(mod.dcic_alias)
                    appendModification(mod.pk, dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["Biosource"][0],singleBio, "modifications",(",".join(modList)))
            
            if(biosource.references):
                appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"references",(biosource.references.dcic_alias))
                appendPublication(biosource.references.pk,dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["Biosource"][0],singleBio,"url",(biosource.url))
            appendFilterdcic(dcicExcelSheet,'Biosource',singleBio)
    
    experiments = experimentList
    
    ##Experiments
    for exp in experiments:
        if(finalizeOnly):
            update_dcic(exp)
        expSet = ExperimentSet.objects.filter(experimentSet_exp=exp)
        if (str(exp.type) == "Hi-C Exp Protocol") or (str(exp.type) == "CaptureC Exp Protocol") or (str(exp.type) =="3C Exp Protocol"):
            if (str(exp.type) == "Hi-C Exp Protocol" or (str(exp.type) =="3C Exp Protocol")):
                singleExp = [""] * len(dcicExcelSheet["ExperimentHiC"][0])
            else:
                singleExp = [""] * len(dcicExcelSheet["ExperimentCaptureC"][0])
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"aliases",(exp.dcic_alias))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"description",(exp.experiment_description))
            experiment_set_join= "" 
            replicate_set_join= ""
            if(expSet):
                experiment_set = []
                replicate_set = []
                for eSet in expSet:
                    if(finalizeOnly):
                        update_dcic(eSet)
                    ExpSet = [""] * len(dcicExcelSheet["ExperimentSet"][0])
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"aliases",(eSet.dcic_alias))
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"description",(eSet.description))
                    appendLab(eSet,"ExperimentSet",ExpSet,dcicExcelSheet) 
                    if(eSet.document):
                        appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"documents",(eSet.document.dcic_alias))
                        appendDocument(eSet.document.pk, dcicExcelSheet,finalizeOnly)
                    if("replicates" in str(eSet.experimentSet_type)):
                        appendFilterdcic(dcicExcelSheet,'ExperimentSetReplicate',ExpSet)
                        replicate_set.append(eSet.dcic_alias)
                    else:
                        appendFilterdcic(dcicExcelSheet,'ExperimentSet',ExpSet)
                        experiment_set.append(eSet.dcic_alias)
            
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"replicate_set",(",".join(replicate_set)))
            #appendBioRep(exp.pk,singleExp)
            #appendTechRep(exp.pk,singleExp)
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"bio_rep_no",(exp.bio_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"tec_rep_no",(exp.tec_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"experiment_set",(",".join(experiment_set)))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"biosample",(exp.experiment_biosample.dcic_alias))
            expFields=json.loads(exp.experiment_fields)
            
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"experiment_type",(expFields["experiment_type"]))
            authDocs=[]
            if(exp.authentication_docs.all()):
                for authDoc in exp.authentication_docs.all():
                    if(finalizeOnly):
                        update_dcic(authDoc)
                    authDocs.append(authDoc.dcic_alias)
                    appendProtocol(authDoc.pk,dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"authentication_docs",(",".join(authDocs)))
            
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"biosample_quantity",(exp.biosample_quantity))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"biosample_quantity_units",(exp.biosample_quantity_units))
            if("biotin_removed" in expFields):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"biotin_removed",(expFields["biotin_removed"]))
            appendLab(exp,"ExperimentHiC",singleExp,dcicExcelSheet) 
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"crosslinking_method",(expFields["crosslinking_method"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"crosslinking_temperature",(expFields["crosslinking_temperature"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"crosslinking_time",(expFields["crosslinking_time"]))
            if(exp.experiment_enzyme):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"digestion_enzyme",(exp.experiment_enzyme.dcic_alias))
                appendEnzyme(exp.experiment_enzyme.pk, dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"digestion_temperature",(expFields["digestion_temperature"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"digestion_time",(expFields["digestion_time"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"enzyme_lot_number",(expFields["enzyme_lot_number"]))
            if("follows_sop" in expFields):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"follows_sop",(expFields["follows_sop"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"average_fragment_size",(expFields["average_fragment_size"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"fragment_size_range",(expFields["fragment_size_range"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"fragment_size_selection_method",(expFields["fragment_size_selection_method"]))
            if("fragmentation_method" in expFields):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"fragmentation_method",(expFields["fragmentation_method"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"library_preparation_date",(expFields["library_preparation_date"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"ligation_temperature",(expFields["ligation_temperature"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"ligation_time",(expFields["ligation_time"]))
            appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"ligation_volume",(expFields["ligation_volume"]))

            if(exp.protocol):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"protocol",(exp.protocol.dcic_alias))
                appendProtocol(exp.protocol.pk, dcicExcelSheet, finalizeOnly)
            
            if(exp.variation):
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"protocol_variation",(exp.variation.dcic_alias))
                appendProtocol(exp.protocol.pk, dcicExcelSheet, finalizeOnly)
            
            
            if (str(exp.type) == "Hi-C Exp Protocol") or (str(exp.type) =="3C Exp Protocol"):
                if("tagging_method" in expFields):
                    appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"tagging_method",(expFields["tagging_method"]))
                if(SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)):
                    files = SeqencingFile.objects.filter(sequencingFile_exp=exp.pk).order_by('sequencingFile_name')
                    if(files.all()):
                        fileList = []
                        for f in files:
                            if(finalizeOnly):
                                update_dcic(f)
                            fileList.append(f.dcic_alias)
                            appendFiles(f.pk,dcicExcelSheet,finalizeOnly)
                        appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"files",(",".join(fileList)))
                
                if(exp.document):
                    appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"documents",(exp.document.dcic_alias))
                    appendDocument(exp.document.pk, dcicExcelSheet,finalizeOnly)
                
                if(exp.references):
                    appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"references",(exp.references.dcic_alias))
                    appendPublication(exp.references.pk, dcicExcelSheet,finalizeOnly)
                    
                appendingFunc(dcicExcelSheet["ExperimentHiC"][0],singleExp,"dbxrefs",(exp.dbxrefs))
                appendFilterdcic(dcicExcelSheet,'ExperimentHiC',singleExp)
            else:
                appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"rna_tag",(expFields["rna_tag"]))
                appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"tagging_method",(expFields["tagging_method"]))
                if(SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)):
                    files = SeqencingFile.objects.filter(sequencingFile_exp=exp.pk).order_by('sequencingFile_name')
                    if(files.all()):
                        fileList = []
                        for f in files:
                            if(finalizeOnly):
                                update_dcic(f)
                            fileList.append(f.dcic_alias)
                            appendFiles(f.pk,dcicExcelSheet,finalizeOnly)
                        appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"files",(",".join(fileList)))
                
                if(exp.document):
                    appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"documents",(exp.document.dcic_alias))
                    appendDocument(exp.document.pk, dcicExcelSheet,finalizeOnly)
                
                if(exp.references):
                    appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"references",(exp.references.dcic_alias))
                    appendPublication(exp.references.pk, dcicExcelSheet,finalizeOnly)
                    
                appendingFunc(dcicExcelSheet["ExperimentCaptureC"][0],singleExp,"dbxrefs",(exp.dbxrefs))
                appendFilterdcic(dcicExcelSheet,'ExperimentCaptureC',singleExp)
            
            
        #   ----------
        elif str(exp.type) == "ATAC-seq Protocol":
            singleExp = [""] * len(dcicExcelSheet["ExperimentAtacseq"][0])
            
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"aliases",(exp.dcic_alias))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"description",(exp.experiment_description))
            if(expSet):
                experiment_set = []
                replicate_set = []
                for eSet in expSet:
                    if(finalizeOnly):
                        update_dcic(eSet)
                    ExpSet = [""] * len(dcicExcelSheet["ExperimentSet"][0])
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"aliases",(eSet.dcic_alias))
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"description",(eSet.description))
                    appendLab(eSet,"ExperimentSet",ExpSet,dcicExcelSheet)
                    if(eSet.document):
                        appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"documents",(eSet.document.dcic_alias))
                        appendDocument(eSet.document.pk, dcicExcelSheet,finalizeOnly)
                    if("replicates" in str(eSet.experimentSet_type)):
                        appendFilterdcic(dcicExcelSheet,'ExperimentSetReplicate',ExpSet)
                        replicate_set.append(eSet.dcic_alias)
                    else:
                        appendFilterdcic(dcicExcelSheet,'ExperimentSet',ExpSet)
                        experiment_set.append(eSet.dcic_alias)
            
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"replicate_set",(",".join(replicate_set)))
            #appendBioRep(exp.pk,singleExp)
            #appendTechRep(exp.pk,singleExp)
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"bio_rep_no",(exp.bio_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"tec_rep_no",(exp.tec_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"experiment_set",(",".join(experiment_set)))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"biosample",(exp.experiment_biosample.dcic_alias))
            
            expFields=json.loads(exp.experiment_fields)
            
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"experiment_type",(expFields["experiment_type"]))
            
            if(exp.authentication_docs.all()):
                for authDoc in exp.authentication_docs.all():
                    if(finalizeOnly):
                        update_dcic(authDoc)
                    authDocs.append(authDoc.dcic_alias)
                    appendProtocol(authDoc.pk,dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"authentication_docs",(",".join(authDocs)))
            
                      
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"biosample_quantity",(exp.biosample_quantity))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"biosample_quantity_units",(exp.biosample_quantity_units))
            appendLab(exp,"ExperimentAtacseq",singleExp,dcicExcelSheet) 
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"enzyme_incubation_time",(expFields["enzyme_incubation_time"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"enzyme_lot_number",(expFields["enzyme_lot_number"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"follows_sop",(expFields["follows_sop"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"incubation_temperature",(expFields["incubation_temperature"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"library_preparation_date",(expFields["library_preparation_date"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"pcr_cycles",(expFields["pcr_cycles"]))
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"primer_removal_method",(expFields["primer_removal_method"]))
            if(exp.protocol):
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"protocol",(exp.protocol.dcic_alias))
                appendProtocol(exp.protocol.pk, dcicExcelSheet, finalizeOnly)
            if(exp.variation):
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"protocol_variation",(exp.variation.dcic_alias))
                appendProtocol(exp.variation.pk,dcicExcelSheet,finalizeOnly)
            if(exp.experiment_enzyme):
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"transposase",(exp.experiment_enzyme.dcic_alias))
                appendEnzyme(exp.experiment_enzyme.pk, dcicExcelSheet,finalizeOnly)
            
            if(SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)):
                files = SeqencingFile.objects.filter(sequencingFile_exp=exp.pk).order_by('sequencingFile_name')
                if(files.all()):
                    fileList = []
                    for f in files:
                        if(finalizeOnly):
                            update_dcic(f)
                        fileList.append(f.dcic_alias)
                        appendFiles(f.pk,dcicExcelSheet,finalizeOnly)
                    appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"files",(",".join(fileList)))
            
            if(exp.document):
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"documents",(exp.document.dcic_alias))
                appendDocument(exp.document.pk, dcicExcelSheet,finalizeOnly)
            
            if(exp.references):
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"documents",(exp.references.dcic_alias))
                appendPublication(exp.references.pk, dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"documents",(exp.dbxrefs))
            appendFilterdcic(dcicExcelSheet,'ExperimentAtacseq',singleExp)
            
            
             #   ----------
        elif str(exp.type) == "ExperimentSeq Protocol":
            singleExp = [""] * len(dcicExcelSheet["ExperimentSeq"][0])
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"aliases",(exp.dcic_alias))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"description",(exp.experiment_description))
            if(expSet):
                experiment_set = []
                replicate_set = []
                for eSet in expSet:
                    if(finalizeOnly):
                        update_dcic(eSet)
                    ExpSet = [""] * len(dcicExcelSheet["ExperimentSet"][0])
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"aliases",(eSet.dcic_alias))
                    appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"description",(eSet.description))
                    appendLab(eSet,"ExperimentSet",ExpSet,dcicExcelSheet)
                    if(eSet.document):
                        appendingFunc(dcicExcelSheet["ExperimentSet"][0],ExpSet,"documents",(eSet.document.dcic_alias))
                        appendDocument(eSet.document.pk, dcicExcelSheet,finalizeOnly)
                    if("replicates" in str(eSet.experimentSet_type)):
                        appendFilterdcic(dcicExcelSheet,'ExperimentSetReplicate',ExpSet)
                        replicate_set.append(eSet.dcic_alias)
                    else:
                        appendFilterdcic(dcicExcelSheet,'ExperimentSet',ExpSet)
                        experiment_set.append(eSet.dcic_alias)
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"replicate_set",(",".join(replicate_set)))
            #appendBioRep(exp.pk,singleExp)
            #appendTechRep(exp.pk,singleExp)
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"bio_rep_no",(exp.bio_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"tec_rep_no",(exp.tec_rep_no))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"experiment_set",(",".join(experiment_set)))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"biosample",(exp.experiment_biosample.dcic_alias))
            
            expFields=json.loads(exp.experiment_fields)
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"experiment_type",(expFields["experiment_type"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"3p_adenylation_temperature",(expFields["3p_adenylation_temperature"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"3p_adenylation_time",(expFields["3p_adenylation_time"]))
            
            if(exp.antibody):
                appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"antibody",(exp.antibody.dcic_alias))
                appendAntibody(exp.antibody.pk, dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"antibody_dilution",(expFields["antibody_dilution"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"antibody_lot_id",(expFields["antibody_lot_id"]))
            
            if(exp.authentication_docs.all()):
                for authDoc in exp.authentication_docs.all():
                    if(finalizeOnly):
                        update_dcic(authDoc)
                    authDocs.append(authDoc.dcic_alias)
                    appendProtocol(authDoc.pk,dcicExcelSheet,finalizeOnly)
                appendingFunc(dcicExcelSheet["ExperimentAtacseq"][0],singleExp,"authentication_docs",(",".join(authDocs)))
            
                      
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"biosample_quantity",(exp.biosample_quantity))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"biosample_quantity_units",(exp.biosample_quantity_units))
            appendLab(exp,"ExperimentSeq",singleExp,dcicExcelSheet)
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"follows_sop",(expFields["follows_sop"]))
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"average_fragment_size",(expFields["average_fragment_size"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"fragment_size_range",(expFields["fragment_size_range"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"fragment_size_selection_method",(expFields["fragment_size_selection_method"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"fragmentation_method",(expFields["fragmentation_method"]))
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"library_preparation_date",(expFields["library_preparation_date"]))
            
            if(exp.protocol):
                appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"protocol",(exp.protocol.dcic_alias))
                appendProtocol(exp.protocol.pk, dcicExcelSheet, finalizeOnly)
            if(exp.variation):
                appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"protocol_variation",(exp.variation.dcic_alias))
                appendProtocol(exp.variation.pk,dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"reaction_time",(expFields["reaction_time"]))
            
            if(SeqencingFile.objects.filter(sequencingFile_exp=exp.pk)):
                files = SeqencingFile.objects.filter(sequencingFile_exp=exp.pk).order_by('sequencingFile_name')
                if(files.all()):
                    fileList = []
                    for f in files:
                        if(finalizeOnly):
                            update_dcic(f)
                        fileList.append(f.dcic_alias)
                        appendFiles(f.pk,dcicExcelSheet,finalizeOnly)
                    appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"files",(",".join(fileList)))
            
            if(exp.document):
                appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"documents",(exp.document.dcic_alias))
                appendDocument(exp.document.pk, dcicExcelSheet,finalizeOnly)
            
            if(exp.references):
                appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"documents",(exp.references.dcic_alias))
                appendPublication(exp.references.pk, dcicExcelSheet,finalizeOnly)
            
            appendingFunc(dcicExcelSheet["ExperimentSeq"][0],singleExp,"documents",(exp.dbxrefs))
            appendFilterdcic(dcicExcelSheet,'ExperimentSeq',singleExp)
            
        # -------
    return(dcicExcelSheet)

def duplicates(lst, item):
    return [i for i, x in enumerate(lst) if x == item]

def removeDup(dcicExcelSheet):
    newdcicExcelSheet = defaultdict(list)
    for key, valueList in dcicExcelSheet.items():
        newValueList = []
        pk = []
        for v in valueList:
            if(v[0] not in pk):
                newValueList .append(v)
                pk.append(v[0])
            else:
                continue
#         pos = dict((x, duplicates(pk, x)) for x in set(pk) if pk.count(x) > 1)    
        newdcicExcelSheet[key] = newValueList
    return(newdcicExcelSheet)

# def initializeSheet(wb):
#     columnNamesDict=defaultdict(list)
#     sheetnames= wb.get_sheet_names()
#     for sheet in sheetnames:
#         v=list(wb.get_sheet_by_name(sheet).rows)[0]
#         for i in v[1:len(v)]:
#             i.value=re.sub('\*', '', str(i.value))
#             columnNamesDict[sheet].append(i.value)
#     return(columnNamesDict)   
                
@login_required 
def exportDCIC(request):
    expPks = request.POST.getlist('dcic')
    if(len(expPks) != 0):
        experiments = Experiment.objects.filter(pk__in=expPks)
    else:
        projectId = request.session['projectId']
        experiments = Experiment.objects.filter(project=projectId)
    if(all(ExperimentSet.objects.filter(experimentSet_exp=exp) for exp in experiments)):
        # Create the HttpResponse object with the appropriate CSV header.
        if(request.session['finalizeOnly']):
            dcicExcelSheet = populateDict(request, experiments)
            for e in experiments:
                e.finalize_dcic_submission=True
                e.save()
            messages.success(request, 'All checked experiments have been marked as DCIC submitted')
            return
        else:
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=DCIC.xlsx'
            file_path_new = ROOTFOLDER+'/organization/static/siteWide/DCIC-Metadata_entry_form_FEB2018.xlsx'
            wb = load_workbook(file_path_new)
            dcicExcelSheet = populateDict(request, experiments)
            
            dcicExcelSheetOrdered = collections.OrderedDict(dcicExcelSheet)
            
            dcicExcelSheetDedup = removeDup(dcicExcelSheetOrdered)
            
            for key, valueList in dcicExcelSheetDedup.items():
                ws = wb.get_sheet_by_name(key)
                for r in reversed(list(ws.rows)):
                    values = [cell.value for cell in r] 
                    if any(values):
                        maxRow=r[0].row+1
                        break
                del valueList[0]
                for v in valueList:
                    for i in range(0,len(v)):
                        ws.cell(row=maxRow, column=i+2).value = v[i]
                    maxRow +=1
            
         
            wb.save(response)
            return response
    else:
        messages.error(request, 'Please add the all experiments into biological/technical replicate experiment set!')
        return HttpResponseRedirect('/detailProject/'+request.session['projectId'])
